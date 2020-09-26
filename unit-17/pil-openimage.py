import PIL as pic
import os, openpyxl
from openpyxl.drawing.image import Image as pic

path = os.path.join(os.getcwd(), "images/工美2")
images = os.listdir(path)
imagesize = 150, 200
files =[]
for image in images:
    if not image.lower().endswith(".jpg"):
        images.remove(image)
    else:
        files.append(os.path.join(path,image))
wb = openpyxl.Workbook()
ws = wb.active
i = 2
ws.column_dimensions['B'].width = 20
for image in images:
    print(image)
    img = pic(os.path.join(path,image))
    img.width = 150
    ws.row_dimensions[i].height = 150
    img.height = 200
    ws.cell(i,1).value = image
    address = "B{}".format(i)
    ws.add_image(img, address)
    i += 1
header = ['文件名',"照片","姓名","身份证号"]
i = 1
for head in header:
    ws.cell(1, i).value = head
    i += 1
wb.save("addimages.xlsx")
wb.close()

