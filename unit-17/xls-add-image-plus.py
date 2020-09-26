import os
import openpyxl
from openpyxl.drawing.image import Image as pic
# step 1 获取指定文件夹下的班级照片
def getImageInfo(path = "images"):
    path = os.path.join(os.getcwd(),"images")
    dirs = os.listdir(path)
    imagesinfo = {}
    for dir in dirs:
        if os.path.isdir(os.path.join(path,dir)):
            print(dir)
            images = []
            files = os.listdir(os.path.join(path, dir))
            for file in files:
                if file.lower().endswith(".jpg"):
                    print(file)
                    images.append(os.path.join(path,dir,file))
            imagesinfo[dir] = images

    return imagesinfo
# step 2 写入excel表格
header=['文件名',"照片","姓名","身份证号"]
wb = openpyxl.Workbook()
imgsize = 150, 200


def setwscom(ws, header=['文件名', "照片", "姓名", "身份证号"]):
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 20
    ws.append(header)


imaggesinfo = getImageInfo()
for key in imaggesinfo.keys():
    wbt = openpyxl.Workbook()
    wst = wbt.active
    wst.title = key
    ws = wb.create_sheet(key)
    setwscom(ws)
    setwscom(wst)
    files = imaggesinfo[key]
    i = 2
    for file in files:
        image = pic(file)
        image.width, image.height = imgsize
        ws.row_dimensions[i].height = 150
        wst.row_dimensions[i].height = 150
        ws.cell(i, 1).value = os.path.basename(file)
        wst.cell(i, 1).value = os.path.basename(file)
        address = "B{}".format(i)
        ws.add_image(image, address)
        wst.add_image(image, address)
        i += 1
    wbt.save(key+".xlsx")
    wbt.close()
wb.save('temp2.xlsx')
wb.close()
