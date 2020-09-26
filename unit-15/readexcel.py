import  openpyxl
workbook = openpyxl.load_workbook("../data/example.xlsx")

for ws in workbook:
    print(type(ws), ws, sep= "    ")
#_________________________________________
### 复制表格
wsc = workbook.active
for row  in range(1,wsc.max_row+1):
    print(row)
    for cell in range(1, wsc.max_column+1):
        if row != 3 :
            continue
        workbook['Sheet2'].cell(row=row, column = cell).value = (wsc.cell(row= row, column= cell).value)
workbook.save("../data/example.xlsx")
workbook.close()