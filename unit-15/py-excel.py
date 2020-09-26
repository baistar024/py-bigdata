# import openpyxl
#
# # read a excel file
# # step 1 open excel file
# workbook = openpyxl.load_workbook("../data/example.xlsx")
# # return <class 'openpyxl.workbook.workbook.Workbook'>
#
# # handle workbook object
# # get_sheet_names
# # sheets = workbook.get_sheet_names()
# #可以在工作表中使用get_highest_row(),get_hight_column()方法确定表的大小
# sheet1 = workbook.get_sheet_by_name("Sheet1")
# print(sheet1)
# print(sheet1.title)
# activesheet = workbook.get_active_sheet()
# print(activesheet.title)
# 访问一个单元格可以使用引用关键字，也可以使用参数.cell(row=m,column=n),使用单元格的值可以使用.value属性,同时cell对象有row、column 和coordinate
# print(sheet1['a1'])
# print(sheet1['a1'].value)
# print(sheet1['b1'].value)
# print(sheet1['b1'].row)
# print(sheet1['b1'].column)
# print(sheet1.cell(row =4, column =5).coordinate)
#列表可以在数字和字母之间黑转换openpyxl.cell.column_index_from_string() 字母->数字
# openpyxl.cell.get_column_letter()   数字->字母

# print(type(sheet1['A1:B3']))
# # print(sheet1['A1:B3'])
# #访问单元区域
# block = sheet1['a1:b4']
# for row in block:
#     for cell in row:
#         print(cell.coordinate, cell.value)
#     print('-'*20)
#
# #工作表中一行一列，也可以使用行号或列标来访问
# line = sheet1.rows[1]
# print(type(line))
# for cell in tuple(line):
#     print(cell.coordinate, cell.value)
# 工作簿、工作表、单元格
# 作为快速复习，下面是从电子表格文件中读取单元格涉及的所有函数、方法和数据类型。
# 1．导入openpyxl 模块。
# 2．调用openpyxl.load_workbook()函数。取得Workbook 对象。
# 3．调用get_active_sheet()或get_sheet_by_name()工作簿方法。取得Worksheet 对象。
# 4．使用索引或工作表的cell()方法，带上row 和column 关键字参数。取得Cell 对象。
# 5．读取Cell 对象的value 属性。

import openpyxl, pprint
workbook = openpyxl.load_workbook("../data/example.xlsx")
worksheet = workbook.get_sheet_by_name("Sheet1")
countsdata = {}
print('Reading rows ...')
for row in range(2, worksheet.get_highest_row() + 1):
    claname = worksheet["C"+str(row)].value
    stuname = worksheet["E"+str(row)].value
    roomnum = worksheet["G"+str(row)].value
    bednum = worksheet["F"+str(row)].value
    print(claname,stuname,roomnum,bednum)
